from numpy import full
from selenium.webdriver import Chrome, ChromeOptions
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time
import openpyxl
import datetime

dt_now=datetime.datetime.now()
dtnf=dt_now.strftime('%Y%m%d_%H%M%S')



def set_driver(hidden_chrome: bool=False):
    USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36"
    options = ChromeOptions()
    options.add_argument('--headless')
    options.add_argument(f'--user-agent={USER_AGENT}') # ブラウザの種類を特定するための文字列
    options.add_argument('log-level=3') # 不要なログを非表示にする
    options.add_argument('--ignore-certificate-errors') # 不要なログを非表示にする
    options.add_argument('--ignore-ssl-errors') # 不要なログを非表示にする
    options.add_experimental_option('excludeSwitches', ['enable-logging']) # 不要なログを非表示にする
    options.add_argument('--incognito') # シークレットモードの設定を付与
    
    service=Service(ChromeDriverManager().install())
    return Chrome(service=service, options=options)


def main():
    
    wb=openpyxl.load_workbook("ツール読込用データ.xlsx")
    wsinfo=wb['検索情報']
    wsresult=wb["検索結果"]
    
    # selenium起動
    driver=set_driver()
    print("実行中です。しばらくお待ちください")
    url="https://www.traders.co.jp/market_jp/screening"
    driver.get(url)
    # 検索結果に無駄な情報が入っていたらクリアする。
    wsresult.delete_rows(2,10000)
    
    full_cell=4
    for i in range(4,51):
        try:
            if wsinfo.cell(i,11).value:
                full_cell+=1
        except:
            pass
    print(f"検索列数は{full_cell-4}件です")
    
    # 検索値を読み込み
    for cnt in range(4,full_cell+1):
        # while True:
            try:
                print(f"{cnt-3}件目開始しました。")
                amount=wsinfo.cell(cnt,11).value
                from_finprc=wsinfo.cell(cnt,12).value
                to_finprc=wsinfo.cell(cnt,13).value
                from_yes=wsinfo.cell(cnt,14).value
                to_yes=wsinfo.cell(cnt,15).value
            # 検索画面で設定値を入力
                time.sleep(2)
                driver.find_element(by=By.CSS_SELECTOR,value=("#flg_sel03")).click()
                driver.find_element(by=By.CSS_SELECTOR,value=("#flg_sel04")).click()
                driver.find_element(by=By.CSS_SELECTOR,value=("#close_price_min")).send_keys(from_finprc)
                driver.find_element(by=By.CSS_SELECTOR,value=("#close_price_max")).send_keys(to_finprc)
                driver.find_element(by=By.CSS_SELECTOR,value=("#change_price_min")).send_keys(from_yes)
                driver.find_element(by=By.CSS_SELECTOR,value=("#change_price_max")).send_keys(to_yes)
                driver.find_element(by=By.CSS_SELECTOR,value=("#submit_btn")).click()
                time.sleep(1)
                
                # 検索結果が０件だったら、なしと記入したい。
                # 検索結果trを読み込み、tdを抽出する。
                try:
                    elem_nodata=driver.find_element(by=By.CSS_SELECTOR,value=".no_data") 
                    max_row=wsresult.max_row
                    if wsinfo.cell(cnt-1,11).value=="分量":
                        wsresult.cell(row=max_row+1,column=1).value=amount
                        wsresult.cell(row=max_row+1,column=2).value="なし"
                    elif wsinfo.cell(cnt,11).value==wsresult.cell(wsresult.max_row,1).value:
                        wsresult.cell(row=max_row+2,column=1).value=amount
                        wsresult.cell(row=max_row+2,column=2).value="なし"
                    else:
                        wsresult.cell(row=max_row+2,column=1).value=amount
                        wsresult.cell(row=max_row+2,column=2).value="なし"
                        
                                    
                except:
                    trs = driver.find_elements(by=By.CSS_SELECTOR, value="table.data_table > tbody > tr") 
                    if trs:
                        for tr in trs:
                            tds=tr.find_elements(by=By.CSS_SELECTOR,value="td")

                            max_row=wsresult.max_row
                            # 「コード」「銘柄」など、ヘッダー行は無視する
                            if len(tds)==0:
                                pass
                            # １回目は、Excelヘッダー行の下に入力させる
                            elif wsinfo.cell(cnt-1,11).value=="分量":
                                for c,td in enumerate(tds):
                                    try:
                                        td=int(td.text)
                                        wsresult.cell(row=max_row+1,column=c+2).value=td               
                                    except:
                                        wsresult.cell(row=max_row+1,column=c+2).value=td.text                
                                wsresult.cell(row=max_row+1,column=1).value=amount
                            # cell(4，11)は、ユーザーが指定した分量が入る。既にwsresult.max_rowに入っている数値と比較して、同じ場合は、maxrowの一つ下に数値を格納。                        
                            elif wsinfo.cell(cnt,11).value==wsresult.cell(wsresult.max_row,1).value:
                                for c,td in enumerate(tds):
                                    try:
                                        td=int(td.text)
                                        wsresult.cell(row=max_row+1,column=c+2).value=td
                                    except:
                                        wsresult.cell(row=max_row+1,column=c+2).value=td.text
                                    # wsresult.cell(row=r+1,column=c+2).value=td.text
                                    wsresult.cell(row=max_row+1,column=1).value=amount
                            else:
                                for c,td in enumerate(tds):
                                    try:
                                        td=int(td.text)
                                        wsresult.cell(row=max_row+2,column=c+2).value=td
                                    except:
                                        wsresult.cell(row=max_row+2,column=c+2).value=td.text
                                                       
                                    wsresult.cell(row=max_row+2,column=1).value=amount
                                    # wsresult.cell(row=r+1,column=c+2).value=td.text
                
                driver.get(url)
                time.sleep(1)
            except:
                print(f"{cnt-3}件目以降データなし")
    
    wb.save("出力結果{}.xlsx".format(dtnf))

main()
'''
dt_now=datetime.datetime.now()
dtnf=dt_now.strftime('%Y%m%d_%H:%M:%S')
'''
'''
elif wsinfo.cell(cnt,11).value==wsresult.cell(wsresult.max_row,1).value:
    for c,td in enumerate(tds):
        wsresult.cell(row=max_row+1,column=c+2).value=td.text                
    wsresult.cell(row=max_row+1,column=1).value=amount
'''    
'''
# ２回目は、１回目の最終行から、１行開けて入力させたい
else:
    for c,td in enumerate(tds):
        # 4，11は、数値が入る。既にresultrowに入っている数値と比較して、同じ場合は、maxrowの一つ下に入れればよい。
        if wsinfo.cell(cnt,11).value==wsresult.cell(wsresult.max_row,1).value:
            wsresult.cell(row=max_row+1,column=c+2).value=td.text                
            # wsresult.cell(row=r+1,column=c+2).value=td.text
            wsresult.cell(row=max_row+1,column=1).value=amount
        else:
            wsresult.cell(row=max_row+2,column=c+2).value=td.text                
        # wsresult.cell(row=r+1,column=c+2).value=td.text
            wsresult.cell(row=max_row+2,column=1).value=amount
'''                
