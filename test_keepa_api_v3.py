import keepa
import json
import openpyxl.drawing.image
import numpy as np
import PIL.Image
import openpyxl
import time
import datetime
import urllib.request

# APIキーを読み込み
file=open('config.json',"r")
j_file=json.load(file)
accesskey=j_file["API_KEY"]
api=keepa.Keepa(accesskey)

# 出力用日付データを取得
dt_now=datetime.datetime.now()
dtnf=dt_now.strftime('%Y%m%d_%H%M%S')

global products
global dst_path
title=""
asins=[]

# Excel読込
wb=openpyxl.load_workbook("asin_to_xl.xlsx") 
ws=wb['general']
url="https://m.media-amazon.com/images/I/"
print("ASINを抽出します。Excelは必ず閉じてください")
time.sleep(1)


def get_asins():
    print("ASINの読取中")
# 変な値が入っていたら値をnoneにする。
    for row in ws.iter_rows(min_row=ws.max_row+1, min_col=1, max_row=5000, max_col=1):
        for cell in row:
            cell.value = None
# ExcelからASINを読み取り、リストasinsに格納
    for row in range(2,ws.max_row+1):
        if ws.cell(row,column=1).value==None:
            pass
        else:
            asins.append(ws.cell(row,column=1).value)
    print("ASINの読取完了")

'''
画像データの読込用関数　本案件では使用しない
def download_file(name):
    dst_path="./{}.png".format(dtnf)
    url=f"https://m.media-amazon.com/images/I/{name}"
    with urllib.request.urlopen(url) as web_file:
        data=web_file.read()
        with open(dst_path,mode="wb") as local_file:
            local_file.write(data) 
def resize_img(dst_path):
    datam=PIL.Image.open(dst_path)
    data_resize=datam.resize((90,90))
    data_resize.save(dst_path)
def paste_toxslx(dst_path,position):
    img_to_excel=openpyxl.drawing.image.Image(dst_path)
    ws.add_image(img_to_excel,position)
'''

def to_excel():
    print("商品情報を呼び出します")
    print(f"商品数量は{len(asins)}件です。")
    
    # keepa apiを叩く。productsにオブジェクト格納。
    products=api.query(asins,domain='JP')
    for i,product in enumerate(products,2):
        
        '''
        商品画像の読み込み用。本案件では使用しない
        if product['imagesCSV']==None:
            print("なし")
        else:
            img_list=product['imagesCSV'].split(",")
            name=img_list[0]
            dst_path="./{}.png".format(dtnf)
            url=f"https://m.media-amazon.com/images/I/{name}"
            with urllib.request.urlopen(url) as web_file:
                data=web_file.read()
                with open(dst_path,mode="wb") as local_file:
                    local_file.write(data) 
            resize_img(dst_path) 
            paste_toxslx(dst_path,f'b{i}')
            wb.save("asin_to_xl.xlsx")
        '''
        
        # JANコード
        if product['eanList']==None:
            ws.cell(i,2).value="なし"
        else:
            ws.cell(i,2).value=int("\n".join(product['eanList']))
            
        # 商品タイトル
        if product['title']==None:
            ws.cell(i,3).value="なし"
        else:
            ws.cell(i,3).value=product['title']
        
        # この商品について
        if product['description']==None:
            ws.cell(i,4).value="なし"
        else:
            pro_des=product['description']
            ws.cell(i,4).value=pro_des.replace("商品紹介      ","")
        
        # 商品の説明
        if product['features']==None:
            ws.cell(i,5).value="なし"
        else:
            ws.cell(i,5).value="\n".join(product['features'])

    print("Excelへの書込完了。ツールを終了します。")
    wb.save("出力結果{}.xlsx".format(dtnf))
    
get_asins()
to_excel()
print(f"残りのトークン数は{api.tokens_left}です。")
time.sleep(2)